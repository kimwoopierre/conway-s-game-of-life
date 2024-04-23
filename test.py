import openpyxl as excel
from openpyxl.styles import PatternFill
wb = excel.load_workbook("qrcode01.xlsx")

sheet = wb["Sheet1"]
data =[]
fill_color = PatternFill(start_color="FFFFFF")

for y in range(1,109):
    for x in range(1,193):
        cell = sheet.cell(row=x, column=y)
        color = cell.fill.start_color.rgb
        if color == None:
            wb.active
            cell.fill = fill_color
wb.save("qrcode01.xlsx")

for y in range(1,109):
    y_data = []
    for x in range(1,193):
        cell = sheet.cell(row=x,column=y)
        color = cell.fill.start_color.rgb
        if color != "00000000":
            y_data.append(True)
        else:
            y_data.append(False)           
        # if color == "FF000000":
        # elif color == "00000000":
    data.append(y_data)

# data = []
# image = []
# image.append((94,40))
# image.append((93,41))
# image.append((93,42))
# image.append((94,42))
# image.append((95,42))



